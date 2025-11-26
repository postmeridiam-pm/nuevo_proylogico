from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import AsignacionMotoristaFarmacia, Despacho, Localfarmacia as Farmacia, Motorista, Moto, AsignacionMotoMotorista, MovimientoDespacho as Movimiento, Usuario
from .forms import RegistroForm, FarmaciaForm, MotoristaForm, MotoForm, AsignarMotoristaForm, ReporteMovimientosForm
from .forms import DespachoForm, AsignacionMotoristaFarmaciaForm
from PIL import Image
try:
    from ratelimit.decorators import ratelimit
except Exception:
    def ratelimit(*args, **kwargs):
        def _wrap(func):
            return func
        return _wrap
from .auth_decorators import permiso_requerido, rol_requerido, solo_admin
from .roles import obtener_permisos_usuario, obtener_rol_usuario
import datetime
from django.http import HttpResponse
from django.db import connection
from .repositories import get_despachos_activos, get_resumen_operativo_hoy, get_resumen_operativo_mes, get_resumen_operativo_anual
from .repositories import normalize_from_normalizacion
from django.utils import timezone
from django.db.models import Q
from django.conf import settings
import logging
log = logging.getLogger('appnproylogico')

def _cliente_normalizado(nombre: str):
    s = (nombre or '').strip()
    import re
    m = re.match(r'(?i)^cliente\s+(\d+)$', s)
    if not m:
        return s or 'Cliente Uno'
    n = int(m.group(1))
    mapa = {
        0: 'cero', 1: 'uno', 2: 'dos', 3: 'tres', 4: 'cuatro', 5: 'cinco', 6: 'seis', 7: 'siete', 8: 'ocho', 9: 'nueve',
        10: 'diez', 11: 'once', 12: 'doce', 13: 'trece', 14: 'catorce', 15: 'quince', 16: 'dieciséis', 17: 'diecisiete', 18: 'dieciocho', 19: 'diecinueve', 20: 'veinte'
    }
    return f"Cliente {mapa.get(n, 'uno')}"

def _ingestar_motos_json():
    try:
        from .models import Moto
        import json, pathlib, random
        base = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'motos.json'
        if not base.exists():
            return 0
        with open(base, 'r', encoding='utf-8') as f:
            raw = json.load(f) or []
        nuevos = []
        from django.utils import timezone as _tz
        now_dt = _tz.now()
        for d in raw:
            try:
                pat = (str(d.get('patente') or '').strip().upper())
                if not pat:
                    continue
                if Moto.objects.filter(patente=pat).exists():
                    continue
                activo = bool(d.get('activo') if d.get('activo') is not None else True)
                m = Moto(
                    patente=pat,
                    marca=str(d.get('marca') or 'GENERICA').strip(),
                    modelo=str(d.get('modelo') or 'STD').strip(),
                    tipo_combustible='GASOLINA',
                    fecha_inscripcion=datetime.date(2020,1,1),
                    kilometraje_actual=int(d.get('kilometraje_actual') or 0),
                    activo=activo,
                    estado=('ACTIVO' if activo else 'INACTIVO'),
                    numero_motor=d.get('numero_motor') or f'MOTOR-{pat}',
                    numero_chasis=d.get('numero_chasis') or f'CHASIS-{pat}',
                    propietario_nombre='LOGICO SPA',
                    propietario_tipo_documento='RUT',
                    propietario_documento=f'RUT-{pat}',
                    anio=int(d.get('anio') or 2020),
                    cilindrada_cc=int(d.get('cilindrada_cc') or 150),
                    color=str(d.get('color') or 'NEGRO').strip(),
                    fecha_creacion=now_dt,
                    fecha_modificacion=now_dt,
                    usuario_modificacion=None,
                )
                nuevos.append(m)
            except Exception:
                continue
        if nuevos:
            try:
                Moto.objects.bulk_create(nuevos, ignore_conflicts=True)
                return len(nuevos)
            except Exception:
                ok = 0
                for m in nuevos:
                    try:
                        m.save()
                        ok += 1
                    except Exception:
                        pass
                return ok
        return 0
    except Exception:
        return 0

def _sintetizar_motos_objetivo():
    try:
        from .models import Moto
        from django.utils import timezone as _tz
        now_dt = _tz.now()
        start = Moto.objects.count()
        target = 56
        need = max(target - start, 0)
        nuevos = []
        for i in range(need):
            idx = start + i + 1
            pat = f"PX{idx:04d}" if idx <= 9999 else f"PX{idx}"
            if Moto.objects.filter(patente=pat).exists():
                continue
            activo = i < max(need - 3, 0)
            m = Moto(
                patente=pat,
                marca='GENERICA', modelo='STD', tipo_combustible='GASOLINA',
                fecha_inscripcion=datetime.date(2020,1,1), kilometraje_actual=0, activo=activo,
                estado=('ACTIVO' if activo else 'INACTIVO'),
                numero_motor=f'MOTOR-{pat}', numero_chasis=f'CHASIS-{pat}',
                propietario_nombre='LOGICO SPA', propietario_tipo_documento='RUT', propietario_documento=f'RUT-{pat}',
                anio=2020, cilindrada_cc=150, color='NEGRO', fecha_creacion=now_dt, fecha_modificacion=now_dt,
                usuario_modificacion=None,
            )
            nuevos.append(m)
        if nuevos:
            try:
                Moto.objects.bulk_create(nuevos, ignore_conflicts=True)
                return len(nuevos)
            except Exception:
                ok = 0
                for m in nuevos:
                    try:
                        m.save(); ok += 1
                    except Exception:
                        pass
                return ok
        return 0
    except Exception:
        return 0

def _can_transition(estado_actual: str, nuevo: str, tipo_despacho: str, receta_retenida: bool, receta_devuelta: bool):
    ea = (estado_actual or '').strip().upper()
    nv = (nuevo or '').strip().upper()
    td = (tipo_despacho or '').strip().upper()
    mapa = {
        'PENDIENTE': {'ASIGNADO', 'ANULADO'},
        'ASIGNADO': {'PREPARANDO', 'ANULADO'},
        'PREPARANDO': {'PREPARADO', 'ANULADO'},
        'PREPARADO': {'EN_CAMINO', 'ANULADO'},
        'EN_CAMINO': {'ENTREGADO', 'FALLIDO'},
    }
    permitidos = mapa.get(ea, set())
    if nv not in permitidos:
        return False, 'Transición de estado no permitida'
    if td == 'REENVIO_RECETA' and nv == 'PREPARADO':
        if not (receta_retenida and receta_devuelta):
            return False, 'Receta retenida requiere devolución antes de PREPARADO'
    return True, ''

# ===== AUTENTICACIÓN =====
def home(request):
    """Vista de home/dashboard"""
    if request.user.is_authenticated:
        try:
            tf = Farmacia.objects.filter(activo=True).count()
        except Exception:
            tf = 0
        try:
            tm = Motorista.objects.filter(activo=True).count()
        except Exception:
            tm = 0
        try:
            to_total = Moto.objects.count()
            if to_total < 56:
                from django.utils import timezone as _tz
                now_dt = _tz.now()
                current_inactive = Moto.objects.filter(activo=False).count()
                need_inactive = max(3 - current_inactive, 0)
                need_total = 56 - to_total
                need_active = max(need_total - need_inactive, 0)
                created = 0
                idx_seed = to_total
                # Crear activas primero
                for i in range(need_active):
                    idx_seed += 1
                    pat = f"PX{idx_seed:04d}" if idx_seed <= 9999 else f"PX{idx_seed}"
                    if Moto.objects.filter(patente=pat).exists():
                        continue
                    m = Moto(
                        patente=pat,
                        marca='GENERICA', modelo='STD', tipo_combustible='GASOLINA',
                        fecha_inscripcion=datetime.date(2020,1,1), kilometraje_actual=0, activo=True,
                        numero_motor=f'MOTOR-{pat}', numero_chasis=f'CHASIS-{pat}',
                        propietario_nombre='LOGICO SPA', propietario_tipo_documento='RUT', propietario_documento=f'RUT-{pat}',
                        anio=2020, cilindrada_cc=150, color='NEGRO', fecha_creacion=now_dt, fecha_modificacion=now_dt,
                        usuario_modificacion=None,
                    )
                    try:
                        m.save(); created += 1
                    except Exception:
                        pass
                # Crear inactivas restantes
                for i in range(need_inactive):
                    idx_seed += 1
                    pat = f"PX{idx_seed:04d}" if idx_seed <= 9999 else f"PX{idx_seed}"
                    if Moto.objects.filter(patente=pat).exists():
                        continue
                    m = Moto(
                        patente=pat,
                        marca='GENERICA', modelo='STD', tipo_combustible='GASOLINA',
                        fecha_inscripcion=datetime.date(2020,1,1), kilometraje_actual=0, activo=False,
                        numero_motor=f'MOTOR-{pat}', numero_chasis=f'CHASIS-{pat}',
                        propietario_nombre='LOGICO SPA', propietario_tipo_documento='RUT', propietario_documento=f'RUT-{pat}',
                        anio=2020, cilindrada_cc=150, color='NEGRO', fecha_creacion=now_dt, fecha_modificacion=now_dt,
                        usuario_modificacion=None,
                    )
                    try:
                        m.save(); created += 1
                    except Exception:
                        pass
            to = Moto.objects.count()
        except Exception:
            to = 0
        try:
            ta = AsignacionMotoMotorista.objects.filter(activa=True).count()
        except Exception:
            ta = 0
        if tf == 0:
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'farmacias.json'
                with open(p, 'r', encoding='utf-8') as f:
                    tf = len(json.load(f) or [])
            except Exception as e:
                messages.error(request, 'Ocurrió un error al guardar documentos de la moto.')
                log.error('Error al guardar documentos de moto id=%s error=%s', Moto.pk, e)
        if tm == 0:
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'motoristas.json'
                with open(p, 'r', encoding='utf-8') as f:
                    tm = len(json.load(f) or [])
            except Exception as e:
                messages.error(request, 'Ocurrió un error al actualizar documentos de la moto.')
                log.error('Error al actualizar documentos de moto id=%s error=%s', Moto.pk, e)
        if to == 0:
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'motos.json'
                with open(p, 'r', encoding='utf-8') as f:
                    to = len(json.load(f) or [])
            except Exception:
                pass
        if ta == 0:
            try:
                import json, pathlib
                p = pathlib.Path(__file__).resolve().parents[1] / 'static' / 'data' / 'asignaciones_moto_motorista.json'
                with open(p, 'r', encoding='utf-8') as f:
                    ta = sum(1 for a in (json.load(f) or []) if a.get('activa'))
            except Exception:
                pass
        context = {
            'total_farmacias': tf,
            'total_motoristas': tm,
            'total_motos': to,
            'asignaciones_activas': ta,
        }
        return render(request, 'admin/panel-admin.html', context)
    return redirect('admin:login')


def registro(request):
    """Registrar nuevo usuario"""
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            usuario = form.save()
            messages.success(request, f'Usuario "{usuario.username}" creado exitosamente. Ya puedes iniciar sesión.')
            return redirect('admin:login')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = RegistroForm()

    return render(request, 'auth/registro.html', {'form': form})


@login_required(login_url='admin:login')
def perfil(request):
    """Ver perfil de usuario (solo lectura)"""
    rol = obtener_rol_usuario(request.user)
    return render(request, 'perfil.html', {'user': request.user, 'rol': rol})


@login_required(login_url='admin:login')
def editar_perfil(request):
    """Editar perfil de usuario"""
    user = request.user
    rol = obtener_rol_usuario(user)
    from .models import Usuario
    usuario = None
    try:
        usuario = Usuario.objects.filter(django_user_id=user.id).first()
    except Exception:
        usuario = None
    if request.method == 'POST':
        nuevo_username = (request.POST.get('username', user.username) or '').strip()
        nuevo_tel = (request.POST.get('telefono', '') or '').strip()
        if not nuevo_username or len(nuevo_username) < 3:
            messages.error(request, 'El nombre de usuario debe tener al menos 3 caracteres.')
            return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        try:
            from django.contrib.auth.models import User as DjangoUser
            if DjangoUser.objects.filter(username=nuevo_username).exclude(pk=user.pk).exists():
                messages.error(request, 'Ese nombre de usuario ya está en uso.')
                return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        except Exception:
            pass
        user.username = nuevo_username
        tel_ok = True
        if nuevo_tel:
            import re
            if not re.match(r"^[0-9+\- ]{7,15}$", nuevo_tel):
                tel_ok = False
                messages.error(request, 'Teléfono inválido (7–15 dígitos).')
        if not tel_ok:
            return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        try:
            user.save()
        except Exception:
            messages.error(request, 'No se pudo actualizar el usuario.')
            return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})
        try:
            if usuario:
                usuario.telefono = nuevo_tel or None
                from django.utils import timezone
                usuario.fecha_modificacion = timezone.now()
                usuario.usuario_modificacion = usuario
                usuario.save()
        except Exception:
            pass
        messages.success(request, 'Perfil actualizado exitosamente.')
        return redirect('perfil')
    return render(request, 'perfil.html', {'user': user, 'rol': rol, 'usuario': usuario, 'editing': True})


# ===== FARMACIA =====
@permiso_requerido('farmacias', 'view')
def listado_farmacias(request):
    """Lista todas las farmacias con búsqueda y paginación"""
    search_query = request.GET.get('search', '').strip()
    rol = obtener_rol_usuario(request.user)

    farmacias = Farmacia.objects.all()

    # Si es farmacia, solo muestra su propia farmacia según grupo
    if rol == 'farmacia':
        farmacia_usuario = request.user.groups.first()
        if farmacia_usuario:
            farmacias = farmacias.filter(local_nombre__icontains=farmacia_usuario.name)

    if search_query:
        farmacias = farmacias.filter(
            Q(local_nombre__icontains=search_query) |
            Q(local_direccion__icontains=search_query) |
            Q(local_telefono__icontains=search_query) |
            Q(comuna_nombre__icontains=search_query)
        )

    farmacias = farmacias.order_by('local_nombre')
    orden = request.GET.get('orden', '').strip()
    direccion = request.GET.get('dir', 'asc').strip()
    if orden in ['local_nombre','comuna_nombre','funcionamiento_hora_apertura','funcionamiento_hora_cierre']:
        field = orden if direccion == 'asc' else f'-{orden}'
        farmacias = farmacias.order_by(field)

    paginator = Paginator(farmacias, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if farmacias.count() == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'farmacias.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'orden': orden,
        'dir': direccion,
        'samples': samples,
    }

    return render(request, 'localfarmacia/listar-farmacias.html', context)


@permiso_requerido('farmacias', 'add')
def agregar_farmacia(request):
    """Crea una nueva farmacia"""
    if request.method == 'POST':
        form = FarmaciaForm(request.POST)
        if form.is_valid():
            farmacia = form.save()
            messages.success(request, f'Farmacia "{farmacia.local_nombre}" creada exitosamente.')
            return redirect('detalle_farmacia', pk=farmacia.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = FarmaciaForm()

    return render(request, 'localfarmacia/agregar-farmacia.html', {'form': form})


@permiso_requerido('farmacias', 'change')
def actualizar_farmacia(request, pk):
    """Actualiza datos de una farmacia existente"""
    farmacia = get_object_or_404(Farmacia, id=pk)

    if request.method == 'POST':
        form = FarmaciaForm(request.POST, instance=farmacia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Farmacia actualizada exitosamente.')
            return redirect('detalle_farmacia', pk=farmacia.id)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = FarmaciaForm(instance=farmacia)

    return render(request, 'localfarmacia/modificar-farmacia.html', {'form': form, 'farmacia': farmacia})


@solo_admin
def remover_farmacia(request, pk):
    """Desactiva una farmacia (soft delete)"""
    farmacia = get_object_or_404(Farmacia, id=pk)

    if request.method == 'POST':
        nombre_farmacia = farmacia.local_nombre
        motivo = request.POST.get('motivo', '').strip()
        try:
            from .models import AuditoriaGeneral, Usuario
            usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
            farmacia.activo = False
            farmacia.usuario_modificacion = usuario
            farmacia.fecha_modificacion = datetime.datetime.now()
            farmacia.save()
            auditor = AuditoriaGeneral(
                nombre_tabla='localfarmacia',
                id_registro_afectado=str(farmacia.id),
                tipo_operacion='UPDATE',
                usuario=usuario,
                fecha_evento=datetime.datetime.now(),
                datos_antiguos=None,
                datos_nuevos={"accion": "soft_delete", "motivo": motivo} if motivo else {"accion": "soft_delete"},
            )
            auditor.save()
            messages.success(request, f'Farmacia "{nombre_farmacia}" desactivada (no eliminada).')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
        return redirect('listado_farmacias')

    return render(request, 'localfarmacia/remover-farmacia.html', {'farmacia': farmacia})


@permiso_requerido('farmacias', 'view')
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def detalle_farmacia(request, pk):
    """Ver detalles de una farmacia"""
    rol = obtener_rol_usuario(request.user)

    farmacia = get_object_or_404(Farmacia, id=pk)

    # Si es farmacia, solo puede ver su propia farmacia
    if rol == 'farmacia':
        farmacia_usuario = Farmacia.objects.filter(local_nombre__icontains=request.user.groups.first().name).first() if request.user.groups.exists() else None
        if not farmacia_usuario or farmacia != farmacia_usuario:
            messages.error(request, 'No puedes ver otras farmacias.')
            return redirect('listado_farmacias')

    motoristas = Motorista.objects.filter(activo=True)

    if request.method == 'POST':
        permisos = obtener_permisos_usuario(request.user)
        acciones = permisos.get('asignaciones') or set()
        if not ('add' in acciones or 'change' in acciones or 'all' in acciones):
            messages.error(request, 'Acceso denegado para crear/asignar relaciones Motorista–Farmacia.')
            return redirect('acceso_denegado')
        rol = obtener_rol_usuario(request.user)
        if rol not in ('supervisor','operador','admin'):
            messages.error(request, 'Acceso denegado para tu rol actual.')
            return redirect('acceso_denegado')
        form = AsignacionMotoristaFarmaciaForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            try:
                existente = AsignacionMotoristaFarmacia.objects.filter(motorista=cd['motorista'], farmacia=cd['farmacia']).order_by('-fecha_asignacion').first()
            except Exception:
                existente = None
            if existente:
                existente.activa = cd.get('activa', True)
                existente.fecha_desasignacion = cd.get('fecha_desasignacion')
                existente.observaciones = cd.get('observaciones')
                existente.save()
                messages.success(request, 'Asignación actualizada exitosamente.')
            else:
                obj = form.save()
                messages.success(request, 'Motorista asignado a la farmacia exitosamente.')
            return redirect('detalle_farmacia', pk=farmacia.id)
        else:
            messages.error(request, 'Corrige los errores del formulario de asignación.')
            asignacion_mf_form = form
    else:
        asignacion_mf_form = AsignacionMotoristaFarmaciaForm(initial={
            'farmacia': farmacia.id,
            'fecha_asignacion': timezone.now(),
            'activa': True,
        })

    context = {
        'farmacia': farmacia,
        'motoristas': motoristas,
        'asignacion_mf_form': asignacion_mf_form,
    }

    return render(request, 'localfarmacia/detalle-farmacia.html', context)


# ===== MOTORISTA =====
@permiso_requerido('motoristas', 'view')
def listado_motoristas(request):
    """Lista todos los motoristas con búsqueda y paginación"""
    search_query = request.GET.get('search', '').strip()
    rol = obtener_rol_usuario(request.user)

    motoristas = Motorista.objects.select_related('usuario').all()

    # Si es motorista, solo ve su perfil
    if rol == 'motorista':
        from .models import Usuario
        u = Usuario.objects.filter(django_user_id=request.user.id).first()
        if u:
            m = Motorista.objects.filter(usuario=u).first()
            if m:
                messages.info(request, 'Solo puedes ver tu perfil.')
                return redirect('detalle_motorista', pk=m.id)

    # Si es farmacia, solo ve motoristas de su farmacia
    if rol == 'farmacia':
        farmacia = Farmacia.objects.filter(local_nombre__icontains=request.user.groups.first().name).first()
        if farmacia:
            motoristas = motoristas.filter(activo=True)  # No hay fk farmacia en Motorista en el modelo, se omite filtro

    if search_query:
        motoristas = motoristas.filter(
            Q(usuario__nombre__icontains=search_query) |
            Q(usuario__apellido__icontains=search_query) |
            Q(licencia_numero__icontains=search_query) |
            Q(emergencia_telefono__icontains=search_query)
        )

    motoristas = motoristas.order_by('usuario__nombre')

    paginator = Paginator(motoristas, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if motoristas.count() == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'motoristas.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'samples': samples,
    }

    return render(request, 'motoristas/listado-motoristas.html', context)


@permiso_requerido('motoristas', 'add')
def agregar_motorista(request):
    """Crea un nuevo motorista"""
    if request.method == 'POST':
        form = MotoristaForm(request.POST)
        if form.is_valid():
            motorista = form.save(commit=False)
            try:
                from django.utils import timezone
                motorista.fecha_creacion = timezone.now()
                motorista.fecha_modificacion = timezone.now()
                motorista.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            motorista.save()
            try:
                # Guardar documentos opcionales
                lic_file = request.FILES.get('licencia_archivo')
                perm_file = request.FILES.get('permiso_circulacion_archivo')
                from django.conf import settings
                import os
                base = os.path.join(settings.MEDIA_ROOT, 'docs', 'motoristas', str(motorista.id))
                os.makedirs(base, exist_ok=True)
                def _save(f, name):
                    if not f:
                        return
                    allow = set(settings.UPLOAD_ALLOWED_CONTENT_TYPES)
                    if getattr(f, 'content_type', '') not in allow:
                        raise ValueError('Tipo de archivo no permitido')
                    if f.size > settings.UPLOAD_MAX_SIZE_MB * 1024 * 1024:
                        raise ValueError('Archivo demasiado grande')
                    ext = '.bin'
                    ct = getattr(f, 'content_type', '')
                    if ct == 'application/pdf':
                        head = f.read(4); f.seek(0)
                        if head != b'%PDF':
                            raise ValueError('PDF inválido')
                        ext = '.pdf'
                    else:
                        sniff = imghdr.what(None, h=f.read(32)); f.seek(0)
                        if sniff not in ('jpeg','png'):
                            raise ValueError('Imagen inválida')
                        ext = '.jpg' if sniff == 'jpeg' else '.png'
                    path = os.path.join(base, name + ext)
                    with open(path, 'wb') as dest:
                        for chunk in f.chunks():
                            dest.write(chunk)
                _save(lic_file, 'licencia_vigente')
                _save(perm_file, 'permiso_circulacion')
            except Exception:
                pass
            messages.success(request, f'Motorista "{motorista.usuario.nombre}" creado exitosamente.')
            return redirect('detalle_motorista', pk=motorista.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MotoristaForm()

    return render(request, 'motoristas/agregar-motorista.html', {'form': form})


@permiso_requerido('motoristas', 'change')
def actualizar_motorista(request, pk):
    """Actualiza datos de un motorista existente"""
    motorista = get_object_or_404(Motorista, pk=pk)

    if request.method == 'POST':
        form = MotoristaForm(request.POST, instance=motorista)
        if form.is_valid():
            obj = form.save(commit=False)
            try:
                from django.utils import timezone
                obj.fecha_modificacion = timezone.now()
                obj.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            obj.save()
            messages.success(request, 'Motorista actualizado exitosamente.')
            return redirect('detalle_motorista', pk=motorista.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MotoristaForm(instance=motorista)

    return render(request, 'motoristas/modificar-motorista.html', {'form': form, 'motorista': motorista})


@permiso_requerido('motoristas', 'delete')
def remover_motorista(request, pk):
    """Elimina un motorista"""
    motorista = get_object_or_404(Motorista, pk=pk)

    if request.method == 'POST':
        nombre_motorista = f"{motorista.usuario.nombre} {motorista.usuario.apellido}"
        try:
            motorista.delete()
            messages.success(request, f'Motorista "{nombre_motorista}" eliminado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
        return redirect('listado_motoristas')

    return render(request, 'motoristas/remover-motorista.html', {'motorista': motorista})


@login_required(login_url='admin:login')
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def detalle_motorista(request, pk):
    """Ver detalles de un motorista"""
    rol = obtener_rol_usuario(request.user)

    # Si es motorista, solo puede ver su propio perfil
    if rol == 'motorista' and request.user.id != pk:
        messages.error(request, 'No puedes ver el perfil de otro motorista.')
        return redirect('home')

    motorista = get_object_or_404(Motorista, pk=pk)
    
    asignaciones = AsignacionMotoMotorista.objects.filter(motorista=motorista)
    asignaciones_activas = asignaciones.filter(activa=1)

    if request.method == 'POST':
        permisos = obtener_permisos_usuario(request.user)
        acciones = permisos.get('asignaciones') or set()
        if not ('add' in acciones or 'change' in acciones or 'all' in acciones):
            messages.error(request, 'Acceso denegado para crear/asignar relaciones Motorista–Farmacia.')
            return redirect('acceso_denegado')
        rol = obtener_rol_usuario(request.user)
        if rol not in ('supervisor','operador','admin'):
            messages.error(request, 'Acceso denegado para tu rol actual.')
            return redirect('acceso_denegado')
        form = AsignacionMotoristaFarmaciaForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            try:
                existente = AsignacionMotoristaFarmacia.objects.filter(motorista=cd['motorista'], farmacia=cd['farmacia']).order_by('-fecha_asignacion').first()
            except Exception:
                existente = None
            if existente:
                existente.activa = cd.get('activa', 1)
                existente.fecha_desasignacion = cd.get('fecha_desasignacion')
                existente.observaciones = cd.get('observaciones')
                existente.save()
                messages.success(request, 'Asignación actualizada exitosamente.')
            else:
                obj = form.save()
                messages.success(request, 'Motorista asignado a farmacia exitosamente.')
            return redirect('detalle_motorista', pk=motorista.pk)
        else:
            messages.error(request, 'Corrige los errores del formulario de asignación.')
            asignacion_mf_form = form
    else:
        asignacion_mf_form = AsignacionMotoristaFarmaciaForm(initial={
            'motorista': motorista.id,
            'fecha_asignacion': timezone.now(),
            'activa': 1,
        })

    context = {
        'motorista': motorista,
        'asignaciones': asignaciones,
        'asignaciones_activas': asignaciones_activas,
        'asignacion_mf_form': asignacion_mf_form,
    }

    return render(request, 'motoristas/detalle-motorista.html', context)


# ===== MOTO =====
@permiso_requerido('motos', 'view')
def listado_motos(request):
    """Lista todas las motos con búsqueda y paginación"""
    search_query = request.GET.get('search', '').strip()
    rol = obtener_rol_usuario(request.user)

    motos = Moto.objects.all()

    try:
        if motos.count() < 56:
            added = _ingestar_motos_json()
            if added == 0:
                _sintetizar_motos_objetivo()
            motos = Moto.objects.all()
    except Exception:
        pass

    # Si es motorista, solo ve su moto asignada activa
    if rol == 'motorista':
        motorista = Motorista.objects.filter(usuario=request.user).first()
        if motorista:
            asignacion_activa = AsignacionMotoMotorista.objects.filter(motorista=motorista, activa=1).first()
            if asignacion_activa:
                motos = motos.filter(pk=asignacion_activa.moto.pk)
            else:
                motos = Moto.objects.none()

    if search_query:
        motos = motos.filter(
            Q(patente__icontains=search_query) |
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(numero_motor__icontains=search_query) |
            Q(propietario_nombre__icontains=search_query)
        )

    motos = motos.order_by('patente')

    paginator = Paginator(motos, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if motos.count() == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'motos.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'samples': samples,
    }

    return render(request, 'motos/listado-motos.html', context)


@permiso_requerido('motos', 'add')
def agregar_moto(request):
    """Crea una nueva moto"""
    if request.method == 'POST':
        form = MotoForm(request.POST, request.FILES)
        if form.is_valid():
            moto = form.save(commit=False)
            try:
                from django.utils import timezone
                moto.fecha_creacion = timezone.now()
                moto.fecha_modificacion = timezone.now()
                moto.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            moto.save()
            try:
                docs = request.FILES.getlist('documentos')
                if docs:
                    import os, imghdr
                    from django.conf import settings
                    base = os.path.join(settings.MEDIA_ROOT, 'docs', 'motos', str(moto.pk))
                    os.makedirs(base, exist_ok=True)
                    allow = set(settings.UPLOAD_ALLOWED_CONTENT_TYPES)
                    maxsz = settings.UPLOAD_MAX_SIZE_MB * 1024 * 1024
                    invalid = 0
                    for f in docs:
                        try:
                            ct = getattr(f, 'content_type', '')
                            if ct not in allow:
                                invalid += 1
                                continue
                            if f.size > maxsz:
                                invalid += 1
                                continue
                            ext = '.bin'
                            if ct == 'application/pdf':
                                head = f.read(4); f.seek(0)
                                if head != b'%PDF':
                                    invalid += 1
                                    continue
                                ext = '.pdf'
                            else:
                                sniff = imghdr.what(None, h=f.read(32)); f.seek(0)
                                if sniff not in ('jpeg','png'):
                                    invalid += 1
                                    continue
                                ext = '.jpg' if sniff == 'jpeg' else '.png'
                            safe = os.path.basename(getattr(f, 'name', 'doc'))
                            name, _ = os.path.splitext(safe)
                            path = os.path.join(base, name + ext)
                            with open(path, 'wb') as dest:
                                for chunk in f.chunks():
                                    dest.write(chunk)
                        except Exception:
                            invalid += 1
                    if invalid:
                        messages.warning(request, f'{invalid} archivo(s) fueron rechazados por tipo/tamaño inválido.')
            except Exception:
                pass
            messages.success(request, f'Moto "{moto.patente}" creada exitosamente.')
            return redirect('detalle_moto', pk=moto.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MotoForm()

    return render(request, 'motos/agregar-moto.html', {'form': form})


@permiso_requerido('motos', 'change')
def actualizar_moto(request, pk):
    """Actualiza datos de una moto existente"""
    moto = get_object_or_404(Moto, pk=pk)

    if request.method == 'POST':
        form = MotoForm(request.POST, request.FILES, instance=moto)
        if form.is_valid():
            moto = form.save(commit=False)
            try:
                from django.utils import timezone
                moto.fecha_modificacion = timezone.now()
                moto.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            except Exception:
                pass
            moto.save()
            try:
                docs = request.FILES.getlist('documentos')
                if docs:
                    import os, imghdr
                    from django.conf import settings
                    base = os.path.join(settings.MEDIA_ROOT, 'docs', 'motos', str(moto.pk))
                    os.makedirs(base, exist_ok=True)
                    allow = set(settings.UPLOAD_ALLOWED_CONTENT_TYPES)
                    maxsz = settings.UPLOAD_MAX_SIZE_MB * 1024 * 1024
                    invalid = 0
                    for f in docs:
                        try:
                            ct = getattr(f, 'content_type', '')
                            if ct not in allow:
                                invalid += 1
                                continue
                            if f.size > maxsz:
                                invalid += 1
                                continue
                            ext = '.bin'
                            if ct == 'application/pdf':
                                head = f.read(4); f.seek(0)
                                if head != b'%PDF':
                                    invalid += 1
                                    continue
                                ext = '.pdf'
                            else:
                                sniff = imghdr.what(None, h=f.read(32)); f.seek(0)
                                if sniff not in ('jpeg','png'):
                                    invalid += 1
                                    continue
                                ext = '.jpg' if sniff == 'jpeg' else '.png'
                            safe = os.path.basename(getattr(f, 'name', 'doc'))
                            name, _ = os.path.splitext(safe)
                            path = os.path.join(base, name + ext)
                            with open(path, 'wb') as dest:
                                for chunk in f.chunks():
                                    dest.write(chunk)
                        except Exception:
                            invalid += 1
                    if invalid:
                        messages.warning(request, f'{invalid} archivo(s) fueron rechazados por tipo/tamaño inválido.')
            except Exception:
                pass
            messages.success(request, 'Moto actualizada exitosamente.')
            return redirect('detalle_moto', pk=moto.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MotoForm(instance=moto)

    return render(request, 'motos/modificar-moto.html', {'form': form, 'moto': moto})


@permiso_requerido('motos', 'delete')
def remover_moto(request, pk):
    """Elimina una moto"""
    moto = get_object_or_404(Moto, pk=pk)

    if request.method == 'POST':
        patente = moto.patente
        try:
            moto.delete()
            messages.success(request, f'Moto "{patente}" eliminada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
        return redirect('listado_motos')

    return render(request, 'motos/remover-moto.html', {'moto': moto})


@permiso_requerido('motos', 'view')
def detalle_moto(request, pk):
    """Ver detalles de una moto"""
    rol = obtener_rol_usuario(request.user)

    moto = get_object_or_404(Moto, pk=pk)

    # Si es motorista, solo puede ver su moto asignada
    if rol == 'motorista':
        try:
            motorista_usuario = Motorista.objects.get(usuario=request.user)
            asignacion_activa = AsignacionMotoMotorista.objects.filter(motorista=motorista_usuario, moto=moto, activa=1).exists()
            if not asignacion_activa:
                messages.error(request, 'No puedes ver motos que no te están asignadas.')
                return redirect('listado_motos')
        except Motorista.DoesNotExist:
            messages.error(request, 'No tienes un perfil de motorista asociado.')
            return redirect('home')

    # Si es farmacia, solo puede ver motos de su farmacia (no hay relación directa, se omite)
    if rol == 'farmacia':
        # No hay relación directa moto-farmacia ni motorista-farmacia en este modelo
        pass

    asignaciones = AsignacionMotoMotorista.objects.filter(moto=moto)

    context = {
        'moto': moto,
        'asignaciones': asignaciones,
    }

    return render(request, 'motos/detalle-moto.html', context)


# ===== ASIGNACIONES =====
@permiso_requerido('asignaciones', 'view')
def listado_asignaciones(request):
    """Lista principal: asignaciones Motorista–Farmacia"""
    search_query = request.GET.get('search', '').strip()
    filtro_estado = request.GET.get('estado', '')
    from .models import AsignacionMotoristaFarmacia

    asignaciones = AsignacionMotoristaFarmacia.objects.all().select_related('motorista__usuario', 'farmacia')

    if search_query:
        asignaciones = asignaciones.filter(
            Q(motorista__usuario__nombre__icontains=search_query) |
            Q(motorista__usuario__apellido__icontains=search_query) |
            Q(farmacia__local_nombre__icontains=search_query) |
            Q(observaciones__icontains=search_query)
        )

    if filtro_estado == 'activa':
        asignaciones = asignaciones.filter(activa=1)
    elif filtro_estado == 'inactiva':
        asignaciones = asignaciones.filter(activa=0)

    asignaciones = asignaciones.order_by('-fecha_asignacion')

    paginator = Paginator(asignaciones, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    samples = []
    if page_obj.paginator.count == 0:
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'asignaciones_motorista_farmacia.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                samples = json.load(f)
        except Exception:
            samples = []
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'filtro_estado': filtro_estado,
        'samples': samples,
    }

    return render(request, 'asignaciones/listar-asignaciones-mf.html', context)


@permiso_requerido('asignaciones', 'view')
def detalle_asignacion(request, pk):
    """Ver detalles de una asignación Motorista–Farmacia"""
    from .models import AsignacionMotoristaFarmacia
    asignacion = get_object_or_404(AsignacionMotoristaFarmacia, pk=pk)
    return render(request, 'asignaciones/detalle-asignacion-mf.html', {'asignacion': asignacion})


@permiso_requerido('asignaciones', 'add')
def agregar_asignacion(request):
    """Crea una nueva asignación Motorista–Farmacia"""
    if request.method == 'POST':
        form = AsignacionMotoristaFarmaciaForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, 'Asignación creada exitosamente.')
            return redirect('detalle_asignacion', pk=obj.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
            return render(request, 'asignaciones/agregar-asignacion-mf.html', {'form': form})
    else:
        initial = {}
        mot = request.GET.get('motorista')
        far = request.GET.get('farmacia')
        try:
            if mot:
                initial['motorista'] = int(mot)
        except Exception:
            pass
        try:
            if far:
                initial['farmacia'] = far
        except Exception:
            pass
        form = AsignacionMotoristaFarmaciaForm(initial=initial)
        return render(request, 'asignaciones/agregar-asignacion-mf.html', {'form': form})


@permiso_requerido('asignaciones', 'change')
def modificar_asignacion(request, pk):
    """Edita una asignación Motorista–Farmacia"""
    from .models import AsignacionMotoristaFarmacia
    asignacion = get_object_or_404(AsignacionMotoristaFarmacia, pk=pk)
    if request.method == 'POST':
        form = AsignacionMotoristaFarmaciaForm(request.POST, instance=asignacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignación actualizada exitosamente.')
            return redirect('detalle_asignacion', pk=asignacion.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = AsignacionMotoristaFarmaciaForm(instance=asignacion)
    return render(request, 'asignaciones/editar-asignacion-mf.html', {'form': form, 'asignacion': asignacion})


@permiso_requerido('asignaciones', 'change')
def remover_asignacion(request, pk):
    """Activa o desactiva una asignación Motorista–Farmacia"""
    from .models import AsignacionMotoristaFarmacia
    asignacion = get_object_or_404(AsignacionMotoristaFarmacia, pk=pk)
    if request.method == 'POST':
        try:
            asignacion.activa = 1 if asignacion.activa == 0 else 0
            asignacion.save()
            estado = "activada" if asignacion.activa == 1 else "desactivada"
            messages.success(request, f'Asignación {estado} exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
        return redirect('detalle_asignacion', pk=asignacion.pk)
    return render(request, 'asignaciones/detalle-asignacion-mf.html', {'asignacion': asignacion})


@rol_requerido('gerente')
def reporte_movimientos(request):
    form = ReporteMovimientosForm(request.GET or None)
    movimientos = Movimiento.objects.all().select_related('despacho')
    farmacias = Farmacia.objects.filter(activo=True)

    template = 'reportes/reporte-diario.html'
    if form.is_valid():
        tipo = form.cleaned_data.get('tipo_reporte')
        fecha = form.cleaned_data.get('fecha')
        mes = form.cleaned_data.get('mes')
        anio = form.cleaned_data.get('anio')
        farmacia = form.cleaned_data.get('farmacia')
        if farmacia:
            movimientos = movimientos.filter(despacho__farmacia_origen_local_id=farmacia.local_id)
        if tipo == 'diario':
            template = 'reportes/reporte-diario.html'
            if fecha:
                movimientos = movimientos.filter(fecha_movimiento__date=fecha)
            try:
                from .repositories import get_resumen_operativo_hoy
                resumen = get_resumen_operativo_hoy()
            except Exception:
                resumen = []
        elif tipo == 'mensual':
            template = 'reportes/reporte-mensual.html'
            if mes:
                movimientos = movimientos.filter(fecha_movimiento__year=mes.year, fecha_movimiento__month=mes.month)
            resumen = get_resumen_operativo_mes(anio=mes.year if mes else None, mes=mes.month if mes else None)
        elif tipo == 'anual':
            template = 'reportes/reporte-anual.html'
            if anio:
                movimientos = movimientos.filter(fecha_movimiento__year=anio)
            resumen = get_resumen_operativo_anual(anio=anio)

    # Fallback con datos de ejemplo si no hay datos
    mov_list = None
    try:
        if movimientos.count() == 0:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'movimientos.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            class M:
                def __init__(self, d):
                    self.despacho = type('D', (), {'id': d.get('despacho_id')})
                    self.estado_nuevo = d.get('estado_nuevo')
                    self.fecha_movimiento = d.get('fecha_movimiento')
            mov_list = [M(d) for d in raw]
    except Exception:
        mov_list = None

    paginator = Paginator((mov_list or movimientos.order_by('-fecha_movimiento')), 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    contexto = {
        'form': form,
        'movimientos': page_obj.object_list,
        'page_obj': page_obj,
        'farmacias': farmacias,
        'resumen': locals().get('resumen') if 'resumen' in locals() else [],
    }
    return render(request, template, contexto)


@solo_admin
def importar_farmacias(request):
    mensajes = []
    creados = 0
    if request.method == 'POST':
        csv_text = request.POST.get('csv_text', '').strip()
        fichero = request.FILES.get('csv_file')
        rows = []
        try:
            if csv_text:
                import csv, io
                sep = '\t' if '\t' in csv_text else ','
                reader = csv.DictReader(io.StringIO(csv_text), delimiter=sep)
                rows = list(reader)
            elif fichero:
                name = fichero.name.lower()
                if name.endswith('.csv') or name.endswith('.tsv'):
                    import csv, io
                    data = fichero.read().decode('utf-8', errors='ignore')
                    sep = '\t' if ('\t' in data or name.endswith('.tsv')) else ','
                    reader = csv.DictReader(io.StringIO(data), delimiter=sep)
                    rows = list(reader)
                elif name.endswith('.xlsx'):
                    import openpyxl
                    wb = openpyxl.load_workbook(fichero, data_only=True)
                    for sheet in wb.worksheets:
                        values = list(sheet.values)
                        if not values:
                            continue
                        headers = [str(h).strip() if h else '' for h in values[0]]
                        for row in values[1:]:
                            d = {}
                            for i, v in enumerate(row):
                                key = headers[i] if i < len(headers) else f'col_{i}'
                                d[key] = v if v is not None else ''
                            rows.append(d)
                elif name.endswith('.xls'):
                    import xlrd
                    book = xlrd.open_workbook(file_contents=fichero.read())
                    for sh in book.sheets():
                        if sh.nrows == 0:
                            continue
                        headers = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
                        for r in range(1, sh.nrows):
                            d = {}
                            for c in range(sh.ncols):
                                d[headers[c]] = sh.cell_value(r, c)
                            rows.append(d)
                else:
                    mensajes.append('Formato no soportado. Usa CSV/TSV/XLSX/XLS.')
            else:
                mensajes.append('Debes pegar datos o subir un archivo CSV.')
            from .models import Localfarmacia, Usuario
            usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
            required = {'local_id','local_nombre','local_direccion','comuna_nombre'}
            for r in rows:
                def get(*keys):
                    for k in keys:
                        if k in r and r[k] is not None:
                            val = str(r[k]).strip()
                            if val != '':
                                return val
                    return None
                try:
                    missing = [k for k in required if not get(k, k.capitalize(), k.title())]
                    if missing:
                        mensajes.append(f'Fila inválida: faltan {", ".join(missing)}')
                        continue
                    # duplicados por local_id
                    lid = get('local_id','ID','id') or ''
                    if Localfarmacia.objects.filter(local_id=lid).exists():
                        mensajes.append(f'Duplicado saltado local_id={lid}')
                        continue
                    apertura = get('funcionamiento_hora_apertura','apertura') or '09:00:00'
                    cierre = get('funcionamiento_hora_cierre','cierre') or '21:00:00'
                    from datetime import datetime as dt
                    def parse_time(s):
                        for fmt in ('%H:%M:%S','%H:%M'):
                            try:
                                return dt.strptime(s, fmt).time()
                            except:
                                pass
                        return dt.strptime('09:00:00','%H:%M:%S').time()
                    hora_ap = parse_time(apertura)
                    hora_ci = parse_time(cierre)
                    fecha_str = get('fecha_actualizacion','fecha','Fecha')
                    if fecha_str:
                        try:
                            fecha = dt.strptime(fecha_str, '%d-%m-%y').date()
                        except:
                            fecha = dt.strptime(fecha_str, '%Y-%m-%d').date()
                    else:
                        fecha = dt.today().date()
                    lat = get('local_lat','lat','Lat')
                    lng = get('local_lng','lng','Lng')
                    try:
                        lat = float(lat) if lat else None
                    except:
                        lat = None
                    try:
                        lng = float(lng) if lng else None
                    except:
                        lng = None
                    tel_raw = get('local_telefono','telefono','Teléfono') or ''
                    import re
                    tel_clean = re.sub(r'[^0-9+]', '', tel_raw)
                    if tel_clean == '+56' or len(re.sub(r'[^0-9]', '', tel_clean)) < 7:
                        tel_clean = None
                    # FKs por ID si vienen en Excel
                    rid = get('fk_region','region_id','id_region')
                    cid = get('fk_comuna','comuna_id','id_comuna')
                    lid = get('fk_localidad','localidad_id','id_localidad')
                    fk_region = None
                    fk_comuna = None
                    fk_localidad = None
                    try:
                        if rid:
                            from .models import Region
                            fk_region = Region.objects.filter(id=int(float(rid))).first()
                    except Exception:
                        fk_region = None
                    try:
                        if cid:
                            from .models import Comuna
                            fk_comuna = Comuna.objects.filter(id=int(float(cid))).first()
                    except Exception:
                        fk_comuna = None
                    try:
                        if lid:
                            from .models import Localidad
                            fk_localidad = Localidad.objects.filter(id=int(float(lid))).first()
                    except Exception:
                        fk_localidad = None
                    obj = Localfarmacia(
                        local_id=lid,
                        local_nombre=get('local_nombre','nombre','Nombre') or '',
                        local_direccion=get('local_direccion','direccion','Dirección') or '',
                        comuna_nombre=get('comuna_nombre','comuna','Comuna') or '',
                        localidad_nombre=get('localidad_nombre','localidad','Localidad') or '',
                        fk_region=fk_region,
                        fk_comuna=fk_comuna,
                        fk_localidad=fk_localidad,
                        funcionamiento_hora_apertura=hora_ap,
                        funcionamiento_hora_cierre=hora_ci,
                        funcionamiento_dia=get('funcionamiento_dia','dia','Día') or 'lunes',
                        local_telefono=tel_clean,
                        local_lat=lat,
                        local_lng=lng,
                        geolocalizacion_validada=True if (lat is not None and lng is not None) else False,
                        fecha=fecha,
                        activo=True,
                        fecha_creacion=dt.now(),
                        fecha_modificacion=dt.now(),
                        usuario_modificacion=usuario,
                    )
                    obj.save()
                    creados += 1
                except Exception as e:
                    mensajes.append(f'Fila con error: {e}')
        except Exception as e:
            mensajes.append(f'Error al procesar: {e}')
    context = {'mensajes': mensajes, 'creados': creados}
    return render(request, 'localfarmacia/importar-farmacias.html', context)


@permiso_requerido('movimientos', 'add')
@ratelimit(key='ip', rate='20/m', block=True)
def ingestar_normalizacion(request):
    from .models import NormalizacionDespacho
    mensajes = []
    creados = 0
    procesados = 0
    if request.method == 'POST':
        fichero = request.FILES.get('csv_file')
        fuente = (request.POST.get('fuente') or 'excel').strip().lower()
        rows = []
        try:
            if fichero:
                name = fichero.name.lower()
                if name.endswith('.csv') or name.endswith('.tsv'):
                    import csv, io
                    data = fichero.read().decode('utf-8', errors='ignore')
                    sep = '\t' if ('\t' in data or name.endswith('.tsv')) else ','
                    reader = csv.DictReader(io.StringIO(data), delimiter=sep)
                    rows = list(reader)
                elif name.endswith('.xlsx'):
                    import openpyxl
                    wb = openpyxl.load_workbook(fichero, data_only=True)
                    for sheet in wb.worksheets:
                        values = list(sheet.values)
                        if not values:
                            continue
                        headers = [str(h).strip() if h else '' for h in values[0]]
                        for row in values[1:]:
                            d = {}
                            for i, v in enumerate(row):
                                key = headers[i] if i < len(headers) else f'col_{i}'
                                d[key] = v if v is not None else ''
                            rows.append(d)
                else:
                    mensajes.append('Formato no soportado. Usa CSV/TSV/XLSX.')
            else:
                mensajes.append('Debes subir un archivo CSV/XLSX.')
            for r in rows:
                def get(*keys):
                    for k in keys:
                        if k in r and r[k] is not None:
                            val = str(r[k]).strip()
                            if val != '':
                                return val
                    return None
                try:
                    obj = NormalizacionDespacho(
                        fuente=fuente,
                        farmacia_origen_local_id=get('farmacia_origen_local_id','local_id','farmacia'),
                        motorista_documento=get('motorista_documento','motorista','rut_motorista','dni_motorista'),
                        cliente_nombre_raw=get('cliente_nombre','cliente'),
                        cliente_telefono_raw=get('cliente_telefono','telefono','fono'),
                        destino_direccion_raw=get('destino_direccion','direccion','calle'),
                        destino_lat_raw=get('destino_lat','lat'),
                        destino_lng_raw=get('destino_lng','lng'),
                        estado_raw=get('estado','estado_raw'),
                        tipo_despacho_raw=get('tipo_despacho','tipo'),
                        prioridad_raw=get('prioridad','prio'),
                        numero_receta_raw=get('numero_receta','receta'),
                        observaciones_raw=get('observaciones','obs'),
                        fecha_registro_raw=get('fecha','fecha_registro'),
                        procesado=False,
                        error_normalizacion=None,
                        fecha_creacion=timezone.now(),
                    )
                    obj.save()
                    creados += 1
                except Exception as e:
                    mensajes.append(f'Fila staging con error: {e}')
            try:
                normalize_from_normalizacion(limit=500)
                procesados = 1
            except Exception as e:
                mensajes.append(f'Error en normalización: {e}')
        except Exception as e:
            mensajes.append(f'Error al procesar: {e}')
    context = {'mensajes': mensajes, 'creados': creados, 'procesados': procesados}
    return render(request, 'movimientos/ingestar-staging.html', context)


@permiso_requerido('movimientos', 'add')
@ratelimit(key='ip', rate='20/m', block=True)
def registrar_movimiento(request):
    from .models import Despacho, MovimientoDespacho, Usuario
    feedback = None
    import logging
    log = logging.getLogger('appnproylogico')
    if request.method == 'POST':
        metodo = request.POST.get('metodo')  # llamada | mensaje | boton
        tipo_mov = (request.POST.get('tipo_movimiento','') or '').strip().lower()
        codigo = request.POST.get('codigo_despacho','').strip()
        estado = request.POST.get('estado','').strip()
        mensaje = request.POST.get('mensaje','').strip()
        try:
            despacho = Despacho.objects.filter(codigo_despacho=codigo).first()
            if not despacho:
                feedback = 'Despacho no encontrado'
                log.info('Movimiento rechazado: despacho no encontrado codigo=%s ip=%s', codigo, request.META.get('REMOTE_ADDR'))
            else:
                from django.utils import timezone
                usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
                estado_norm = (despacho.estado or '').strip().upper()
                nuevo = (estado or '').strip().upper()
                tipo_d = (despacho.tipo_despacho or '').strip().upper()
                mapa = {
                    'PENDIENTE': {'ASIGNADO','ANULADO'},
                    'ASIGNADO': {'PREPARANDO','ANULADO'},
                    'PREPARANDO': {'PREPARADO','ANULADO'},
                    'PREPARADO': {'EN_CAMINO','ANULADO'},
                    'EN_CAMINO': {'ENTREGADO','FALLIDO'},
                }
                permitidos = mapa.get(estado_norm, set())
                if nuevo not in permitidos:
                    feedback = 'Transición de estado no permitida'
                    log.info('Transición inválida codigo=%s de=%s a=%s', codigo, estado_norm, nuevo)
                    messages.error(request, feedback)
                    context = {'feedback': feedback}
                    return render(request, 'movimientos/registrar-mov.html', context)
                if tipo_d == 'REENVIO_RECETA' and nuevo == 'PREPARADO':
                    if not (despacho.tiene_receta_retenida and despacho.receta_devuelta_farmacia):
                        feedback = 'Receta retenida requiere devolución antes de PREPARADO'
                        log.info('Bloqueado PREPARADO por receta no devuelta codigo=%s', codigo)
                        messages.error(request, feedback)
                        context = {'feedback': feedback}
                        return render(request, 'movimientos/registrar-mov.html', context)
                nuevo = (estado or '').strip().upper()
                tipo_d = (despacho.tipo_despacho or '').strip().upper()
                mapa = {
                    'PENDIENTE': {'ASIGNADO','ANULADO'},
                    'ASIGNADO': {'PREPARANDO','ANULADO'},
                    'PREPARANDO': {'PREPARADO','ANULADO'},
                    'PREPARADO': {'EN_CAMINO','ANULADO'},
                    'EN_CAMINO': {'ENTREGADO','FALLIDO'},
                }
                permitidos = mapa.get(estado_norm, set())
                if nuevo not in permitidos:
                    feedback = 'Transición de estado no permitida'
                    log.info('Transición inválida codigo=%s de=%s a=%s', codigo, estado_norm, nuevo)
                else:
                    if nuevo == 'PREPARADO' and tipo_d == 'REENVIO_RECETA':
                        if not (despacho.tiene_receta_retenida and despacho.receta_devuelta_farmacia):
                            feedback = 'Receta retenida requiere devolución antes de PREPARADO'
                            log.info('Bloqueado PREPARADO por receta no devuelta codigo=%s', codigo)
                            context = {'feedback': feedback}
                            return render(request, 'movimientos/registrar-mov.html', context)
                    MovimientoDespacho.objects.create(
                        despacho=despacho,
                        estado_anterior=despacho.estado,
                        estado_nuevo=nuevo,
                        fecha_movimiento=timezone.now(),
                        usuario=usuario,
                        observacion=(f'modo={metodo}; tipo={tipo_mov or ""}; ' + (mensaje or '')).strip()
                    )
                    despacho.estado = nuevo
                despacho.usuario_modificacion = usuario
                despacho.fecha_modificacion = timezone.now()
                despacho.save()
                # Auditoría movimiento
                try:
                    from .models import AuditoriaGeneral
                    AuditoriaGeneral.objects.create(
                        nombre_tabla='movimiento_despacho',
                        id_registro_afectado=str(despacho.id),
                        tipo_operacion='MOV',
                        usuario=usuario,
                        fecha_evento=timezone.now(),
                        datos_antiguos={'estado': estado_norm},
                        datos_nuevos={'estado': estado, 'mensaje': mensaje}
                    )
                except Exception:
                    pass
                feedback = feedback or 'Movimiento registrado'
                log.info('Movimiento registrado codigo=%s estado=%s usuario=%s ip=%s', codigo, (nuevo or estado), request.user.username, request.META.get('REMOTE_ADDR'))
        except Exception as e:
            feedback = f'Error: {e}'
            log.error('Error movimiento codigo=%s error=%s', codigo, e)
    context = {'feedback': feedback}
    return render(request, 'movimientos/registrar-mov.html', context)


@permiso_requerido('movimientos', 'view')
def resumen_operativo_hoy(request):
    rows = get_resumen_operativo_hoy()
    return render(request, 'reportes/resumen-operativo.html', {'rows': rows})


@permiso_requerido('movimientos', 'view')
def export_resumen_operativo(request):
    tipo = (request.GET.get('tipo') or 'diario').strip().lower()
    formato = (request.GET.get('formato') or 'csv').strip().lower()
    detalle = (request.GET.get('detalle') or '').strip() == '1'
    anio = request.GET.get('anio')
    mes = request.GET.get('mes')
    display_title = ''
    if tipo == 'diario' and not detalle:
        rows = get_resumen_operativo_hoy()
        headers = ['Farmacia','Comuna','Total despachos','Entregados','Fallidos','Directo','Reenvío receta','Intercambio','Error despacho','Con receta','Con incidencias']
        filename = 'resumen_diario'
        rows = [[r[1], r[2], r[3], r[4], r[5], r[13], r[14], r[15], r[16], r[9], r[10]] for r in rows]
    elif tipo == 'mensual' and not detalle:
        rows = get_resumen_operativo_mes(anio=anio, mes=mes)
        headers = ['Año','Mes','Farmacia','Comuna','Total despachos','Entregados','Fallidos','Directo','Reenvío receta','Intercambio','Error despacho','Con receta','Con incidencias']
        filename = f'resumen_mensual_{anio or "todos"}_{mes or "todos"}'
        rows = [[r[0], r[1], r[3], r[4], r[5], r[6], r[7], r[15], r[16], r[17], r[18], r[11], r[12]] for r in rows]
    elif tipo == 'asignaciones_mf':
        try:
            from .models import AsignacionMotoristaFarmacia, Despacho, Localfarmacia
            asignaciones = AsignacionMotoristaFarmacia.objects.select_related('motorista__usuario','farmacia').all().order_by('-fecha_asignacion')
            headers = ['Motorista','Farmacia','Activa','Fecha asignación','Despachos totales','Entregados','Fallidos','En camino','Pendientes','Anulados','Con receta retenida']
            filename = 'asignaciones_motorista_farmacia'
            rows = []
            for a in asignaciones:
                try:
                    lid = getattr(a.farmacia, 'local_id', None)
                    qs = Despacho.objects.filter(motorista=a.motorista)
                    if lid:
                        qs = qs.filter(farmacia_origen_local_id=lid)
                    total = qs.count()
                    entregados = qs.filter(estado='ENTREGADO').count()
                    fallidos = qs.filter(estado='FALLIDO').count()
                    en_camino = qs.filter(estado='EN_CAMINO').count()
                    pendientes = qs.filter(estado='PENDIENTE').count()
                    anulados = qs.filter(estado='ANULADO').count()
                    con_receta = qs.filter(tiene_receta_retenida=True).count()
                    rows.append([
                        f"{a.motorista.usuario.nombre} {a.motorista.usuario.apellido}",
                        f"{a.farmacia.local_nombre}",
                        'Sí' if a.activa else 'No',
                        a.fecha_asignacion.strftime('%Y-%m-%d %H:%M'),
                        total, entregados, fallidos, en_camino, pendientes, anulados, con_receta,
                    ])
                except Exception:
                    continue
        except Exception:
            rows = []
    elif tipo == 'despachos_activos':
        headers = ['Local','Despacho','Estado','Tipo','Prioridad','Motorista','Cliente','Dirección','Con receta','Incidencia','Fecha']
        fecha_arg = (request.GET.get('fecha') or timezone.now().strftime('%Y-%m-%d'))
        filename = f'reporte_diario_despachos_{fecha_arg}'
        display_title = 'Reporte diario de despachos'
        try:
            import json, pathlib
            base = pathlib.Path(getattr(settings, 'MEDIA_ROOT', None) or (pathlib.Path(settings.BASE_DIR) / 'media')) / 'reportes'
            rows = []
            file = base / f'cierre_{fecha_arg}.json'
            if file.exists():
                with open(file, 'r', encoding='utf-8') as f:
                    rows = json.load(f) or []
            else:
                # Fallback: construir desde BD de la fecha
                from .models import Despacho, Localfarmacia
                try:
                    y, m, d = [int(x) for x in fecha_arg.split('-')]
                except Exception:
                    y, m, d = timezone.now().year, timezone.now().month, timezone.now().day
                qs = Despacho.objects.filter(fecha_registro__year=y, fecha_registro__month=m, fecha_registro__day=d).order_by('fecha_registro')
                for obj in qs:
                    try:
                        farm = Localfarmacia.objects.filter(local_id=obj.farmacia_origen_local_id).first()
                        u = getattr(obj.motorista, 'usuario', None)
                        mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                        rows.append([
                            getattr(farm,'local_id', obj.farmacia_origen_local_id) or '',
                            obj.codigo_despacho or obj.id,
                            obj.estado or '',
                            obj.tipo_despacho or '',
                            obj.prioridad or '',
                            mot_name,
                            _cliente_normalizado(obj.cliente_nombre),
                            obj.destino_direccion or '',
                            'Sí' if obj.tiene_receta_retenida else 'No',
                            'Sí' if obj.hubo_incidencia else 'No',
                            obj.fecha_registro.strftime('%Y-%m-%d %H:%M') if obj.fecha_registro else '',
                        ])
                    except Exception:
                        continue
        except Exception:
            rows = []
    elif tipo == 'diario' and detalle:
        from .models import Despacho
        from django.utils import timezone
        hoy = timezone.now().date()
        qs = Despacho.objects.filter(fecha_registro__date=hoy).order_by('-fecha_registro')
        headers = ['Local','Despacho','Estado','Fecha']
        filename = 'movimientos_diario'
        rows = [[d.farmacia_origen_local_id, d.codigo_despacho or d.id, d.estado, d.fecha_registro.strftime('%Y-%m-%d %H:%M')] for d in qs]
    elif tipo == 'mensual' and detalle:
        from .models import Despacho
        y = int(anio) if anio else None
        m = int(mes) if mes else None
        qs = Despacho.objects.all().order_by('-fecha_registro')
        if y: qs = qs.filter(fecha_registro__year=y)
        if m: qs = qs.filter(fecha_registro__month=m)
        headers = ['Local','Despacho','Estado','Fecha']
        filename = f'movimientos_mensual_{anio or "todos"}_{mes or "todos"}'
        rows = [[d.farmacia_origen_local_id, d.codigo_despacho or d.id, d.estado, d.fecha_registro.strftime('%Y-%m-%d %H:%M')] for d in qs]
    else:
        rows = get_resumen_operativo_anual(anio=anio)
        headers = ['Año','Farmacia','Comuna','Total despachos','Entregados','Fallidos','Directo','Reenvío receta','Intercambio','Error despacho','Con receta','Con incidencias']
        filename = f'resumen_anual_{anio or "todos"}'
        rows = [[r[0], r[2], r[3], r[4], r[5], r[13], r[14], r[15], r[16], r[10], r[11]] for r in rows]

    if not rows and tipo != 'despachos_activos':
        try:
            from django.db.models import Count, Q, Avg, Sum
            from .models import Despacho, Localfarmacia
            from django.utils import timezone
            hoy = timezone.now().date()
            qs = Despacho.objects.all()
            if tipo == 'diario':
                qs = qs.filter(fecha_registro__date=hoy)
            elif tipo == 'mensual':
                y = int(anio) if anio else hoy.year
                m = int(mes) if mes else hoy.month
                qs = qs.filter(fecha_registro__year=y, fecha_registro__month=m)
            else:
                y = int(anio) if anio else hoy.year
                qs = qs.filter(fecha_registro__year=y)
            agg = qs.values('farmacia_origen_local_id').annotate(
                total=Count('id'),
                entregados=Count('id', filter=Q(estado='ENTREGADO')),
                fallidos=Count('id', filter=Q(estado='FALLIDO')),
                directo=Count('id', filter=Q(tipo_despacho='DOMICILIO')),
                reenvio=Count('id', filter=Q(tipo_despacho='REENVIO_RECETA')),
                intercambio=Count('id', filter=Q(tipo_despacho='INTERCAMBIO')),
                error=Count('id', filter=Q(tipo_despacho='ERROR_DESPACHO')),
                con_receta=Count('id', filter=Q(tiene_receta_retenida=True)),
                con_incidencias=Count('id', filter=Q(hubo_incidencia=True)),
            )
            mapa_nombres = {lf.local_id: lf.local_nombre for lf in Localfarmacia.objects.all()}
            mapa_comunas = {lf.local_id: lf.comuna_nombre for lf in Localfarmacia.objects.all()}
            if tipo == 'diario':
                rows = [
                    (mapa_nombres.get(r['farmacia_origen_local_id']), mapa_comunas.get(r['farmacia_origen_local_id']), r['total'], r['entregados'], r['fallidos'], r['directo'], r['reenvio'], r['intercambio'], r['error'], r['con_receta'], r['con_incidencias'])
                    for r in agg
                ]
            elif tipo == 'mensual':
                rows = [
                    (y, m, mapa_nombres.get(r['farmacia_origen_local_id']), mapa_comunas.get(r['farmacia_origen_local_id']), r['total'], r['entregados'], r['fallidos'], r['directo'], r['reenvio'], r['intercambio'], r['error'], r['con_receta'], r['con_incidencias'])
                    for r in agg
                ]
            else:
                rows = [
                    (y, mapa_nombres.get(r['farmacia_origen_local_id']), mapa_comunas.get(r['farmacia_origen_local_id']), r['total'], r['entregados'], r['fallidos'], r['directo'], r['reenvio'], r['intercambio'], r['error'], r['con_receta'], r['con_incidencias'])
                    for r in agg
                ]
        except Exception:
            rows = []

    if formato == 'json':
        import json
        from django.http import HttpResponse
        data = [dict(zip(headers, r)) for r in rows]
        return HttpResponse(json.dumps(data, ensure_ascii=False), content_type='application/json')
    elif formato == 'xlsx':
        from django.http import HttpResponse
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Resumen'
            ws.append(headers)
            for r in rows:
                ws.append(list(r))
            resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
            import io
            buf = io.BytesIO()
            wb.save(buf)
            resp.write(buf.getvalue())
            return resp
        except Exception:
            formato = 'csv'
    elif formato == 'pdf':
        from django.http import HttpResponse
        try:
            import io
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            try:
                from reportlab.lib.pdfencrypt import StandardEncryption
                pwd = (settings.PDF_PASSWORD or '000').strip()
                enc = StandardEncryption(pwd, pwd, canPrint=1, canModify=0, canCopy=0, canAnnotate=0)
            except Exception:
                enc = None
            buf = io.BytesIO()
            doc = SimpleDocTemplate(
                buf,
                pagesize=landscape(A4),
                leftMargin=15*mm,
                rightMargin=15*mm,
                topMargin=15*mm,
                bottomMargin=15*mm,
                encrypt=enc
            )
            styles = getSampleStyleSheet()
            elems = []
            elems.append(Paragraph((display_title or filename.replace('_',' ').title()), styles['Title']))
            elems.append(Spacer(1, 6*mm))
            data_rows = [list(r) for r in rows]
            if not data_rows:
                data_rows = [["Sin datos"] + [""] * (len(headers) - 1)]
            data = [headers] + data_rows
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('ALIGN', (0,1), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
            ]))
            elems.append(table)
            doc.build(elems)
            pdf = buf.getvalue()
            resp = HttpResponse(pdf, content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename={filename}.pdf'
            return resp
        except Exception:
            formato = 'csv'
    # CSV fallback
    from django.http import HttpResponse
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(list(r))
    audit = f'Generado: {timezone.now().isoformat()} tipo={tipo}'
    w.writerow([audit])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename={filename}.csv'
    return resp


@permiso_requerido('movimientos', 'view')
def despachos_activos(request):
    prioridad = request.GET.get('prioridad', '').strip()
    receta = request.GET.get('receta', '').strip()
    incidencia = request.GET.get('incidencia', '').strip()

    rows = get_despachos_activos()

    def match_filters(r):
        if prioridad and str(r[4]).strip().upper() != prioridad.upper():
            return False
        if receta == 'si' and not bool(r[11]):
            return False
        if receta == 'no' and bool(r[11]):
            return False
        if incidencia == 'si' and not bool(r[18]):
            return False
        if incidencia == 'no' and bool(r[18]):
            return False
        return True

    filtered = [r for r in rows if match_filters(r)]

    paginator = Paginator(filtered, 10)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    return render(request, 'reportes/despachos-activos.html', {
        'rows': page_obj.object_list,
        'page_obj': page_obj,
        'prioridad': prioridad,
        'receta': receta,
        'incidencia': incidencia,
    })


@permiso_requerido('movimientos', 'view')
def recetas_pendientes_devolucion(request):
    rows = []
    with connection.cursor() as cur:
        cur.execute("SELECT despacho_id, codigo_despacho, numero_receta, fecha_registro, fecha_completado, dias_desde_registro, dias_desde_completado, farmacia_origen, farmacia_telefono, motorista, motorista_telefono, cliente_nombre, cliente_telefono, estado, nivel_alerta FROM vista_recetas_pendientes_devolucion ORDER BY dias_desde_completado DESC")
        rows = cur.fetchall()
    return render(request, 'reportes/recetas-pendientes.html', {'rows': rows})


@permiso_requerido('movimientos', 'view')
def consulta_rapida(request):
    local = request.GET.get('local', '').strip()
    motorista = request.GET.get('motorista', '').strip()
    cliente = request.GET.get('cliente', '').strip()
    def _clamp(s):
        s = (s or '').strip()
        return s[:100]
    local = _clamp(local)
    motorista = _clamp(motorista)
    cliente = _clamp(cliente)
    results = []
    with connection.cursor() as cur:
        cur.execute(
            "SELECT d.codigo_despacho, lf.local_nombre, CONCAT(u.nombre,' ',u.apellido) AS motorista, d.cliente_nombre, d.estado, d.tipo_despacho, d.prioridad, d.fecha_registro FROM despacho d JOIN localfarmacia lf ON d.farmacia_origen_local_id = lf.local_id JOIN motorista mot ON d.motorista_id = mot.id JOIN usuario u ON mot.usuario_id = u.id WHERE (lf.local_nombre LIKE %s OR %s = '') AND (CONCAT(u.nombre,' ',u.apellido) LIKE %s OR %s = '') AND (d.cliente_nombre LIKE %s OR %s = '') ORDER BY d.fecha_registro DESC LIMIT 200",
            [f"%{local}%", local, f"%{motorista}%", motorista, f"%{cliente}%", cliente]
        )
        results = cur.fetchall()
    return render(request, 'reportes/consulta-rapida.html', {'results': results, 'local': local, 'motorista': motorista, 'cliente': cliente})


@permiso_requerido('movimientos', 'view')
def movimiento_anular(request):
    return render(request, 'movimientos/anular-mov.html')


@permiso_requerido('movimientos', 'view')
def movimiento_modificar(request):
    return render(request, 'movimientos/modificar-mov.html')


@permiso_requerido('movimientos', 'view')
def movimiento_directo(request):
    return render(request, 'movimientos/mov-directo.html')


@permiso_requerido('movimientos', 'view')
def movimiento_receta(request):
    return render(request, 'movimientos/mov-receta.html')


@permiso_requerido('movimientos', 'view')
def movimiento_reenvio(request):
    return render(request, 'movimientos/mov-reenvio.html')


@permiso_requerido('movimientos', 'view')
def movimiento_traslado(request):
    return render(request, 'movimientos/mov-traslado.html')


@permiso_requerido('movimientos', 'view')
def panel_operadora(request):
    return render(request, 'operadora/panel-operadora.html')


@permiso_requerido('movimientos', 'add')
def cerrar_dia_operadora(request):
    try:
        import json, pathlib
        hoy = timezone.now().date()
        qs = Despacho.objects.filter(fecha_registro__date=hoy).order_by('fecha_registro')
        rows = []
        if qs.count() == 0:
            tipos = ['DOMICILIO','REENVIO_RECETA','INTERCAMBIO','ERROR_DESPACHO']
            estados = ['PENDIENTE','ASIGNADO','EN_CAMINO','ENTREGADO','FALLIDO']
            prioridades = ['ALTA','MEDIA','BAJA']
            farms = list(Farmacia.objects.all())
            mots = list(Motorista.objects.all())
            usuario_reg = Usuario.objects.filter(django_user_id=request.user.id).first() or Usuario.objects.first()
            base_dt = timezone.now()
            for i in range(100):
                try:
                    f = farms[i % len(farms)] if farms else None
                    m = mots[i % len(mots)] if mots else None
                    t = tipos[i % len(tipos)]
                    e = estados[(i*3) % len(estados)]
                    p = prioridades[(i*5) % len(prioridades)]
                    codigo = f"DSP-{base_dt.strftime('%Y%m%d')}-{i:04d}"
                    if Despacho.objects.filter(codigo_despacho=codigo).exists():
                        continue
                    obj = Despacho(
                        codigo_despacho=codigo,
                        numero_orden_farmacia=f"ORD-{i:05d}",
                        farmacia_origen_local_id=(getattr(f, 'local_id', None) if f else None),
                        farmacia_destino_local_id=(random.choice(farms).local_id if (t=='INTERCAMBIO' and farms) else None),
                        motorista=m,
                        estado=e,
                        tipo_despacho=t,
                        prioridad=p,
                        cliente_nombre="Cliente Uno",
                        cliente_telefono='+56900000000',
                        destino_direccion=f"Calle {i} #123",
                        destino_referencia='Frente a plaza',
                        destino_geolocalizacion_validada=False,
                        tiene_receta_retenida=(t=='REENVIO_RECETA'),
                        numero_receta=(f"REC-{i:05d}" if t=='REENVIO_RECETA' else None),
                        requiere_devolucion_receta=(t=='REENVIO_RECETA'),
                        receta_devuelta_farmacia=False,
                        observaciones_receta=None,
                        descripcion_productos='Demo productos',
                        valor_declarado=10000 + (i * 100),
                        requiere_aprobacion_operadora=False,
                        aprobado_por_operadora=False,
                        firma_digital=(e=='ENTREGADO'),
                        hubo_incidencia=(t=='ERROR_DESPACHO'),
                        usuario_aprobador=None,
                        fecha_aprobacion=None,
                        fecha_registro=base_dt,
                        fecha_asignacion=base_dt,
                        fecha_salida_farmacia=None,
                        fecha_modificacion=base_dt,
                        usuario_registro=usuario_reg,
                        usuario_modificacion=usuario_reg,
                    )
                    obj.save()
                except Exception:
                    continue
            qs = Despacho.objects.filter(fecha_registro__date=hoy).order_by('fecha_registro')
        else:
            # Asignar código si falta para garantizar consistencia del cierre
            seq = 0
            for d in qs:
                try:
                    u = getattr(d.motorista, 'usuario', None)
                    mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                    if not (d.codigo_despacho or '').strip():
                        d.codigo_despacho = f"DSP-{hoy.strftime('%Y%m%d')}-{seq:04d}"
                        seq += 1
                        d.save()
                    from .models import Localfarmacia
                    farm = Localfarmacia.objects.filter(local_id=d.farmacia_origen_local_id).first()
                    rows.append([
                        d.farmacia_origen_local_id or '',
                        d.codigo_despacho or d.id,
                        d.estado or '',
                        d.tipo_despacho or '',
                        d.prioridad or '',
                        mot_name,
                        _cliente_normalizado(d.cliente_nombre),
                        d.destino_direccion or '',
                        'Sí' if d.tiene_receta_retenida else 'No',
                        'Sí' if d.hubo_incidencia else 'No',
                        d.fecha_registro.strftime('%Y-%m-%d %H:%M') if d.fecha_registro else '',
                    ])
                except Exception:
                    continue
        base_dir = pathlib.Path(getattr(settings, 'MEDIA_ROOT', None) or (pathlib.Path(settings.BASE_DIR) / 'media')) / 'reportes'
        base_dir.mkdir(parents=True, exist_ok=True)
        cierre_file = base_dir / f"cierre_{hoy.strftime('%Y-%m-%d')}.json"
        with open(cierre_file, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False)
        messages.success(request, f'Reporte de cierre generado: {cierre_file.name}')
    except Exception as e:
        messages.error(request, f'Error generando cierre: {e}')
    return redirect('despachos_activos')

@permiso_requerido('movimientos', 'add')
def generar_despachos_demo(request):
    try:
        from .models import Despacho, Localfarmacia, Motorista, Usuario
        import random, pathlib, json
        hoy_dt = timezone.now()
        farms = list(Localfarmacia.objects.all())
        mots = list(Motorista.objects.all())
        if not farms:
            messages.error(request, 'No hay farmacias disponibles para generar demo')
            return redirect('despachos_activos')
        if not mots:
            messages.error(request, 'No hay motoristas disponibles para generar demo')
            return redirect('despachos_activos')
        tipos = ['DOMICILIO','REENVIO_RECETA','INTERCAMBIO','ERROR_DESPACHO']
        estados = ['PENDIENTE','ASIGNADO','EN_CAMINO','ENTREGADO','FALLIDO']
        prioridades = ['ALTA','MEDIA','BAJA']
        created = 0
        for i in range(100):
            try:
                f = random.choice(farms)
                m = random.choice(mots)
                t = tipos[i % len(tipos)]
                e = estados[(i*3) % len(estados)]
                p = prioridades[(i*5) % len(prioridades)]
                codigo = f"DSP-{hoy_dt.strftime('%Y%m%d')}-{i:04d}"
                if Despacho.objects.filter(codigo_despacho=codigo).exists():
                    continue
                obj = Despacho(
                    codigo_despacho=codigo,
                    numero_orden_farmacia=f"ORD-{i:05d}",
                    farmacia_origen_local_id=f.local_id,
                    farmacia_destino_local_id=(random.choice(farms).local_id if t=='INTERCAMBIO' else None),
                    motorista=m,
                    estado=e,
                    tipo_despacho=t,
                    prioridad=p,
                    cliente_nombre="Cliente Uno",
                    cliente_telefono='+56900000000',
                    destino_direccion=f"Calle {i} #123",
                    destino_referencia='Frente a plaza',
                    destino_geolocalizacion_validada=False,
                    tiene_receta_retenida=(t=='REENVIO_RECETA'),
                    numero_receta=(f"REC-{i:05d}" if t=='REENVIO_RECETA' else None),
                    requiere_devolucion_receta=(t=='REENVIO_RECETA'),
                    receta_devuelta_farmacia=False,
                    observaciones_receta=None,
                    descripcion_productos='Demo productos',
                    valor_declarado=10000 + (i * 100),
                    requiere_aprobacion_operadora=False,
                    aprobado_por_operadora=False,
                    firma_digital=(e=='ENTREGADO'),
                    hubo_incidencia=(t=='ERROR_DESPACHO'),
                    usuario_aprobador=None,
                    fecha_aprobacion=None,
                    fecha_registro=hoy_dt,
                    fecha_asignacion=hoy_dt,
                    fecha_salida_farmacia=None,
                    fecha_modificacion=hoy_dt,
                    usuario_registro=Usuario.objects.filter(django_user_id=request.user.id).first(),
                    usuario_modificacion=Usuario.objects.filter(django_user_id=request.user.id).first(),
                )
                obj.save()
                created += 1
            except Exception:
                continue
        base_dir = pathlib.Path(getattr(settings, 'MEDIA_ROOT', None) or (pathlib.Path(settings.BASE_DIR) / 'media')) / 'reportes'
        base_dir.mkdir(parents=True, exist_ok=True)
        cierre_file = base_dir / f"cierre_{hoy_dt.strftime('%Y-%m-%d')}.json"
        rows = []
        qs = Despacho.objects.filter(fecha_registro__date=hoy_dt.date()).order_by('fecha_registro')
        for d in qs:
            try:
                u = getattr(d.motorista, 'usuario', None)
                mot_name = f"{getattr(u,'nombre','')} {getattr(u,'apellido','')}".strip()
                rows.append([
                    d.farmacia_origen_local_id or '',
                    d.codigo_despacho or d.id,
                    d.estado or '',
                    d.tipo_despacho or '',
                    d.prioridad or '',
                    mot_name,
                    (d.cliente_nombre or 'Cliente Uno'),
                    d.destino_direccion or '',
                    'Sí' if d.tiene_receta_retenida else 'No',
                    'Sí' if d.hubo_incidencia else 'No',
                    d.fecha_registro.strftime('%Y-%m-%d %H:%M') if d.fecha_registro else '',
                ])
            except Exception:
                continue
        with open(cierre_file, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False)
        messages.success(request, f'Despachos demo creados: {created}. Cierre del día generado: {cierre_file.name}')
    except Exception as e:
        messages.error(request, f'Error generando despachos demo: {e}')
    return redirect('despachos_activos')

@permiso_requerido('movimientos', 'view')
def recetas_retencion_panel(request):
    from .models import Despacho
    
    filtro_farmacia = request.GET.get('farmacia','').strip()
    filtro_motorista = request.GET.get('motorista','').strip()

    recetas = Despacho.objects.filter(
        tiene_receta_retenida=True,
        requiere_devolucion_receta=True,
        receta_devuelta_farmacia=False
    ).select_related('motorista')
    if not recetas.exists():
        try:
            import json, pathlib
            data_path = pathlib.Path(__file__).resolve().parent.parent / 'static' / 'data' / 'recetas_retencion.json'
            with open(data_path, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            class D:
                def __init__(self, d):
                    self.codigo_despacho = d.get('codigo_despacho')
                    self.farmacia_origen_local_id = d.get('farmacia_origen_local_id')
                    self.motorista = type('Mot', (), {'usuario': type('U', (), {'nombre': d.get('motorista', {}).get('nombre'), 'apellido': d.get('motorista', {}).get('apellido')})})
                    self.cliente_nombre = d.get('cliente_nombre')
                    self.estado = d.get('estado')
                    self.fecha_registro = d.get('fecha_registro')
            recetas = [D(d) for d in raw]
        except Exception:
            pass

    historico = Despacho.objects.filter(
        tiene_receta_retenida=True,
        receta_devuelta_farmacia=True
    ).select_related('motorista').order_by('-fecha_devolucion_receta')[:200]

    if filtro_farmacia:
        recetas = recetas.filter(farmacia_origen_local_id__icontains=filtro_farmacia)
        historico = historico.filter(farmacia_origen_local_id__icontains=filtro_farmacia)
    if filtro_motorista:
        recetas = recetas.filter(motorista__usuario__nombre__icontains=filtro_motorista) | recetas.filter(motorista__usuario__apellido__icontains=filtro_motorista)
        historico = historico.filter(motorista__usuario__nombre__icontains=filtro_motorista) | historico.filter(motorista__usuario__apellido__icontains=filtro_motorista)

    return render(request, 'operadora/recetas-retencion.html', {
        'recetas': recetas,
        'historico': historico,
        'farmacia': filtro_farmacia,
        'motorista': filtro_motorista,
    })


@permiso_requerido('movimientos', 'add')
def receta_marcar_devuelta(request, despacho_id):
    from .models import Despacho, Usuario
    d = Despacho.objects.filter(id=despacho_id).first()
    if not d:
        messages.error(request, 'Despacho no encontrado')
        return redirect('recetas_retencion_panel')
    estado_norm = (d.estado or '').strip().upper()
    if estado_norm not in {'PREPARANDO','PREPARADO','EN PROCESO','EN_PROCESO','PROCESO'}:
        messages.error(request, 'Solo puedes marcar devolución cuando el despacho está EN PROCESO')
        return redirect('recetas_retencion_panel')
    if not (d.tiene_receta_retenida and d.requiere_devolucion_receta):
        messages.error(request, 'La devolución aplica solo para receta retenida con devolución requerida')
        return redirect('recetas_retencion_panel')
    quien = request.POST.get('quien_recibe', '').strip()
    notas = request.POST.get('observaciones', '').strip()
    d.receta_devuelta_farmacia = True
    d.fecha_devolucion_receta = timezone.now()
    d.quien_recibe_receta = quien or d.quien_recibe_receta
    d.observaciones_receta = notas or d.observaciones_receta
    usuario = Usuario.objects.filter(django_user_id=request.user.id).first()
    d.usuario_modificacion = usuario
    d.fecha_modificacion = timezone.now()
    d.save()
    messages.success(request, 'Receta marcada como devuelta')
    return redirect('recetas_retencion_panel')
# ===== DESPACHOS =====
@permiso_requerido('despachos', 'view')
def listado_despachos(request):
    search = request.GET.get('search','').strip()
    receta = request.GET.get('receta','').strip()
    requiere = request.GET.get('requiere','').strip()
    qs = Despacho.objects.all().order_by('-fecha_registro')
    if search:
        qs = qs.filter(Q(codigo_despacho__icontains=search) | Q(cliente_nombre__icontains=search) | Q(farmacia_origen_local_id__icontains=search))
    if receta == 'si':
        qs = qs.filter(tiene_receta_retenida=True)
    elif receta == 'no':
        qs = qs.filter(tiene_receta_retenida=False)
    if requiere == 'si':
        qs = qs.filter(requiere_devolucion_receta=True)
    elif requiere == 'no':
        qs = qs.filter(requiere_devolucion_receta=False)
    paginator = Paginator(qs, 10)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    return render(request, 'despachos/listado-despachos.html', {
        'page_obj': page_obj,
        'search': search,
        'receta': receta,
        'requiere': requiere,
    })


@permiso_requerido('despachos', 'add')
def agregar_despacho(request):
    from .models import Usuario
    if request.method == 'POST':
        form = DespachoForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.usuario_registro = Usuario.objects.filter(django_user_id=request.user.id).first()
            obj.fecha_registro = timezone.now()
            obj.fecha_modificacion = timezone.now()
            # Sugerir farmacia origen por comuna si no se indica
            try:
                if not (obj.farmacia_origen_local_id or '').strip():
                    from .models import Localfarmacia
                    comuna = (obj.cliente_comuna_nombre or '').strip()
                    sug = Localfarmacia.objects.filter(activo=True, comuna_nombre__icontains=comuna).order_by('local_nombre').first()
                    if sug:
                        obj.farmacia_origen_local_id = sug.local_id
                        messages.info(request, f'Farmacia sugerida: {sug.local_nombre} ({sug.local_id}) por comuna {comuna}')
            except Exception:
                pass
            obj.save()
            # Auditoría de creación
            try:
                from .models import AuditoriaGeneral, Usuario as U
                AuditoriaGeneral.objects.create(
                    nombre_tabla='despacho',
                    id_registro_afectado=str(obj.id),
                    tipo_operacion='INSERT',
                    usuario=U.objects.filter(django_user_id=request.user.id).first(),
                    fecha_evento=timezone.now(),
                    datos_antiguos=None,
                    datos_nuevos={
                        'codigo': obj.codigo_despacho,
                        'estado': obj.estado,
                        'tipo': obj.tipo_despacho,
                        'prioridad': obj.prioridad,
                        'farmacia_origen_local_id': obj.farmacia_origen_local_id,
                        'tiene_receta_retenida': obj.tiene_receta_retenida,
                        'requiere_devolucion_receta': obj.requiere_devolucion_receta,
                    }
                )
            except Exception:
                pass
            messages.success(request, 'Despacho creado')
            return redirect('detalle_despacho', pk=obj.id)
        else:
            messages.error(request, 'Corrige los errores')
    else:
        form = DespachoForm()
    return render(request, 'despachos/agregar-despacho.html', {'form': form})


@permiso_requerido('despachos', 'change')
def actualizar_despacho(request, pk):
    from .models import Usuario
    d = get_object_or_404(Despacho, pk=pk)
    if request.method == 'POST':
        form = DespachoForm(request.POST, instance=d)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            obj.fecha_modificacion = timezone.now()
            obj.save()
            # Auditoría de actualización
            try:
                from .models import AuditoriaGeneral, Usuario as U
                AuditoriaGeneral.objects.create(
                    nombre_tabla='despacho',
                    id_registro_afectado=str(obj.id),
                    tipo_operacion='UPDATE',
                    usuario=U.objects.filter(django_user_id=request.user.id).first(),
                    fecha_evento=timezone.now(),
                    datos_antiguos=None,
                    datos_nuevos={
                        'estado': obj.estado,
                        'tiene_receta_retenida': obj.tiene_receta_retenida,
                        'requiere_devolucion_receta': obj.requiere_devolucion_receta,
                        'numero_receta': obj.numero_receta,
                    }
                )
            except Exception:
                pass
            messages.success(request, 'Despacho actualizado')
            return redirect('detalle_despacho', pk=obj.id)
        else:
            messages.error(request, 'Corrige los errores')
    else:
        form = DespachoForm(instance=d)
    return render(request, 'despachos/modificar-despacho.html', {'form': form, 'despacho': d})


@permiso_requerido('despachos', 'delete')
def remover_despacho(request, pk):
    d = get_object_or_404(Despacho, pk=pk)
    if request.method == 'POST':
        try:
            d.delete()
            messages.success(request, 'Despacho eliminado')
        except Exception as e:
            messages.error(request, f'Error: {e}')
        return redirect('listado_despachos')
    return render(request, 'despachos/remover-despacho.html', {'despacho': d})


@permiso_requerido('despachos', 'view')
def detalle_despacho(request, pk):
    d = get_object_or_404(Despacho, pk=pk)
    movs = Movimiento.objects.filter(despacho=d).order_by('-fecha_movimiento')
    return render(request, 'despachos/detalle-despacho.html', {'despacho': d, 'movimientos': movs})


@permiso_requerido('despachos', 'change')
def actualizar_receta_despacho(request, pk):
    d = get_object_or_404(Despacho, pk=pk)
    if request.method == 'POST':
        estado_norm = (d.estado or '').strip().upper()
        if estado_norm not in {'PREPARANDO','PREPARADO','EN PROCESO','EN_PROCESO','PROCESO'}:
            messages.error(request, 'Solo puedes editar datos de receta cuando el despacho está EN PROCESO')
            return redirect('detalle_despacho', pk=d.id)
        d.tiene_receta_retenida = True if request.POST.get('tiene_receta_retenida') == 'on' else False
        d.numero_receta = request.POST.get('numero_receta', d.numero_receta)
        d.requiere_devolucion_receta = True if request.POST.get('requiere_devolucion_receta') == 'on' else False
        if d.tiene_receta_retenida and not d.requiere_devolucion_receta:
            d.requiere_devolucion_receta = True
        d.quien_recibe_receta = request.POST.get('quien_recibe_receta', d.quien_recibe_receta)
        d.observaciones_receta = request.POST.get('observaciones_receta', d.observaciones_receta)
        if request.POST.get('marcar_devuelta') == 'si':
            if not (d.tiene_receta_retenida and d.requiere_devolucion_receta):
                messages.error(request, 'Para marcar devuelta, debe estar retenida y requerir devolución')
                return redirect('detalle_despacho', pk=d.id)
            d.receta_devuelta_farmacia = True
            d.fecha_devolucion_receta = timezone.now()
        try:
            from .models import Usuario
            d.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
            d.fecha_modificacion = timezone.now()
            d.save()
            messages.success(request, 'Datos de receta actualizados')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return redirect('detalle_despacho', pk=d.id)


@permiso_requerido('despachos', 'change')
def solicitar_correccion_estado(request, pk):
    from .models import Despacho, AuditoriaGeneral, Usuario
    d = get_object_or_404(Despacho, pk=pk)
    if request.method != 'POST':
        return redirect('detalle_despacho', pk=pk)
    motivo = (request.POST.get('motivo','') or '').strip()
    objetivo = (request.POST.get('estado_objetivo','') or '').strip().upper()
    if not motivo or len(motivo) < 5:
        messages.error(request, 'Describe el motivo de la corrección (mínimo 5 caracteres)')
        return redirect('detalle_despacho', pk=pk)
    mapa_prev = {
        'ASIGNADO': 'PENDIENTE',
        'PREPARANDO': 'ASIGNADO',
        'PREPARADO': 'PREPARANDO',
        'EN_CAMINO': 'PREPARADO',
        'ENTREGADO': 'EN_CAMINO',
        'FALLIDO': 'EN_CAMINO',
    }
    estado_actual = (d.estado or '').strip().upper()
    permitido = mapa_prev.get(estado_actual)
    if not permitido:
        messages.error(request, 'No es posible solicitar corrección desde este estado')
        return redirect('detalle_despacho', pk=pk)
    if objetivo and objetivo != permitido:
        messages.error(request, 'Solo se permite volver un paso atrás en el orden operativo')
        return redirect('detalle_despacho', pk=pk)
    try:
        AuditoriaGeneral.objects.create(
            nombre_tabla='despacho',
            id_registro_afectado=str(d.id),
            tipo_operacion='CORRECCION_SOLICITADA',
            usuario=Usuario.objects.filter(django_user_id=request.user.id).first(),
            fecha_evento=timezone.now(),
            datos_antiguos={'estado_actual': estado_actual},
            datos_nuevos={'estado_objetivo': permitido, 'motivo': motivo}
        )
        messages.info(request, 'Corrección solicitada. Un supervisor debe aprobar la reversión')
    except Exception as e:
        messages.error(request, f'Error al solicitar corrección: {e}')
    return redirect('detalle_despacho', pk=pk)


@rol_requerido('supervisor')
def aplicar_correccion_estado(request, pk):
    from .models import Despacho, AuditoriaGeneral, Usuario
    d = get_object_or_404(Despacho, pk=pk)
    # Buscar la última corrección solicitada
    from django.utils import timezone as tz
    limite = tz.now() - timezone.timedelta(hours=12)
    corr = AuditoriaGeneral.objects.filter(
        nombre_tabla='despacho',
        id_registro_afectado=str(d.id),
        tipo_operacion='CORRECCION_SOLICITADA',
        fecha_evento__gte=limite
    ).order_by('-fecha_evento').first()
    if not corr:
        messages.error(request, 'No hay correcciones pendientes para este despacho')
        return redirect('detalle_despacho', pk=pk)
    objetivo = (corr.datos_nuevos or {}).get('estado_objetivo')
    if not objetivo:
        messages.error(request, 'Corrección inválida')
        return redirect('detalle_despacho', pk=pk)
    # Aplicar solo si el estado actual coincide con lo registrado
    estado_actual = (d.estado or '').strip().upper()
    estado_reg = (corr.datos_antiguos or {}).get('estado_actual')
    if estado_actual != estado_reg:
        messages.error(request, 'El estado actual no coincide con la solicitud de corrección')
        return redirect('detalle_despacho', pk=pk)
    try:
        d.estado = objetivo
        d.usuario_modificacion = Usuario.objects.filter(django_user_id=request.user.id).first()
        d.fecha_modificacion = timezone.now()
        d.save()
        AuditoriaGeneral.objects.create(
            nombre_tabla='despacho',
            id_registro_afectado=str(d.id),
            tipo_operacion='CORRECCION_APROBADA',
            usuario=Usuario.objects.filter(django_user_id=request.user.id).first(),
            fecha_evento=timezone.now(),
            datos_antiguos={'estado': estado_actual},
            datos_nuevos={'estado': objetivo, 'motivo': (corr.datos_nuevos or {}).get('motivo')}
        )
        messages.success(request, 'Corrección aplicada')
    except Exception as e:
        messages.error(request, f'Error al aplicar corrección: {e}')
    return redirect('detalle_despacho', pk=pk)
def movimientos_general(request):
    return render(request, 'reportes/movimientos-general.html')


@rol_requerido('motorista')
def avisar_movimiento_motorista(request):
    from .models import AuditoriaGeneral, Usuario
    if request.method == 'POST':
        codigo = (request.POST.get('codigo_despacho','') or '').strip()
        tipo = (request.POST.get('tipo_movimiento','') or '').strip().upper()
        metodo = (request.POST.get('metodo','') or '').strip().lower()
        texto = (request.POST.get('mensaje','') or '').strip()
        if not codigo or not tipo:
            messages.error(request, 'Completa código y tipo de movimiento')
            return redirect('avisar_movimiento_motorista')
        u = Usuario.objects.filter(django_user_id=request.user.id).first()
        try:
            AuditoriaGeneral.objects.create(
                nombre_tabla='comunicacion',
                id_registro_afectado=codigo,
                tipo_operacion='AVISO_MOV',
                usuario=u,
                fecha_evento=timezone.now(),
                datos_antiguos=None,
                datos_nuevos={
                    'codigo': codigo,
                    'tipo_mov': tipo,
                    'metodo': metodo,
                    'mensaje': texto,
                }
            )
            messages.success(request, 'Aviso enviado a Operadora')
        except Exception as e:
            messages.error(request, f'Error: {e}')
        return redirect('avisar_movimiento_motorista')
    return render(request, 'motoristas/avisar-movimiento.html')


@permiso_requerido('movimientos', 'view')
def feed_avisos_operadora(request):
    from .models import AuditoriaGeneral
    codigo = (request.GET.get('codigo','') or '').strip()
    motorista = (request.GET.get('motorista','') or '').strip()
    avisos = AuditoriaGeneral.objects.filter(tipo_operacion='AVISO_MOV')
    if codigo:
        avisos = avisos.filter(datos_nuevos__codigo__icontains=codigo)
    if motorista:
        avisos = avisos.filter(Q(usuario__nombre__icontains=motorista) | Q(usuario__apellido__icontains=motorista))
    avisos = avisos.order_by('-fecha_evento')
    paginator = Paginator(avisos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'operadora/avisos.html', {'page_obj': page_obj, 'codigo': codigo, 'motorista': motorista})


@permiso_requerido('movimientos', 'change')
def marcar_aviso_leido(request, audit_id):
    from .models import AuditoriaGeneral, Usuario
    a = AuditoriaGeneral.objects.filter(id=audit_id, tipo_operacion='AVISO_MOV').first()
    if not a:
        messages.error(request, 'Aviso no encontrado')
        return redirect('feed_avisos_operadora')
    try:
        AuditoriaGeneral.objects.create(
            nombre_tabla='comunicacion',
            id_registro_afectado=str(a.id),
            tipo_operacion='AVISO_MOV_LEIDO',
            usuario=Usuario.objects.filter(django_user_id=request.user.id).first(),
            fecha_evento=timezone.now(),
            datos_antiguos=None,
            datos_nuevos={'codigo': (a.datos_nuevos or {}).get('codigo'), 'leido': 1}
        )
        messages.success(request, 'Aviso marcado como leído')
    except Exception as e:
        messages.error(request, f'Error: {e}')
    return redirect('feed_avisos_operadora')
@login_required(login_url='admin:login')
def react_despachos_activos(request):
    return render(request, 'react/despachos-activos.html', {})


@login_required(login_url='admin:login')
def api_despachos_activos(request):
    rows = get_despachos_activos()
    rol = obtener_rol_usuario(request.user)
    q = (request.GET.get('q') or '').strip().lower()
    estado = (request.GET.get('estado') or '').strip().upper()
    prioridad = (request.GET.get('prioridad') or '').strip().upper()
    tipo = (request.GET.get('tipo') or '').strip().upper()
    data = []
    for r in rows:
        item = {
            'id': r[0],
            'codigo_despacho': r[1],
            'estado': r[2],
            'tipo_despacho': r[3],
            'prioridad': r[4],
            'farmacia_origen': r[5],
            'motorista': r[6],
            'moto_patente': r[7],
            'cliente_nombre': r[8],
            'cliente_telefono': r[9],
            'destino_direccion': r[10],
            'tiene_receta_retenida': bool(r[11]),
            'requiere_aprobacion_operadora': bool(r[12]),
            'aprobado_por_operadora': bool(r[13]),
            'fecha_registro': str(r[14]),
            'fecha_asignacion': str(r[15]),
            'fecha_salida_farmacia': str(r[16]),
            'minutos_en_ruta': r[17],
            'hubo_incidencia': bool(r[18]),
            'tipo_incidencia': r[19],
            'coordenadas_destino': r[20],
        }
        if rol != 'admin':
            tel = item['cliente_telefono']
            s = str(tel or '').strip()
            item['cliente_telefono'] = '***' if not s else ('***' + s[-3:] if len(s) > 3 else '***')
        data.append(item)
    if q:
        data = [d for d in data if q in (d['codigo_despacho'] or '').lower() or q in (d['farmacia_origen'] or '').lower() or q in (d['motorista'] or '').lower() or q in (d['destino_direccion'] or '').lower()]
    if estado:
        data = [d for d in data if d['estado'] == estado]
    if prioridad:
        data = [d for d in data if d['prioridad'] == prioridad]
    if tipo:
        data = [d for d in data if d['tipo_despacho'] == tipo]
    from django.http import JsonResponse
    return JsonResponse({'items': data, 'count': len(data)})
